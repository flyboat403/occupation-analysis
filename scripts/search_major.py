#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
专业教学标准检索脚本

数据源优先级（更新）：
1. 主数据源：assets/moe_pdfs_final.json（JSON索引）
   - 找到专业且有PDF URL → 下载并解析PDF获取完整内容
   - 找到专业但无PDF URL → 返回基本信息
2. 备用数据源：assets/moe_pdfs_md/{类代码}*.md
   - 仅当在moe_pdfs_final.json中未找到专业时使用

流程：
1. 从 moe_pdfs_final.json 查找专业 → 得到 PDF URL
2. 如果有 PDF URL，下载并解析 PDF
3. 如果 moe_pdfs_final.json 中没有，则从 moe_pdfs_md 目录解析
"""

import json
import re
import argparse
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from major_catalog_mapper import get_catalog, get_category_from_major_code, get_md_file_path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def load_major_data(json_path: Path) -> List[Dict]:
    """
    加载专业教学标准JSON数据
    
    Args:
        json_path: JSON文件路径
        
    Returns:
        专业数据列表
        
    Raises:
        FileNotFoundError: 文件不存在
        json.JSONDecodeError: JSON解析失败
    """
    if not json_path.exists():
        raise FileNotFoundError(f"数据文件不存在: {json_path}")
    
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except json.JSONDecodeError as e:
        raise json.JSONDecodeError(f"JSON解析失败: {json_path}", e.doc, e.pos)

def search_by_code(data: List[Dict], code: str) -> Optional[Dict]:
    """按专业代码检索"""
    for item in data:
        if item.get('major_code') == code:
            return item
    return None

def search_by_name(data: List[Dict], name: str) -> List[Dict]:
    """按专业名称检索（支持模糊匹配）"""
    results = []
    for item in data:
        if name in item.get('major_name', ''):
            results.append(item)
    return results

def search_by_level(data: List[Dict], level: str) -> List[Dict]:
    """按教育层次筛选"""
    results = []
    
    level_aliases = {
        '中职': ['中等职业教育', '中职学校', '中职教育', '职高', '中专', '技校'],
        '高职': ['高等职业教育', '高职专科', '专科','高职院校'],
        '职教本科': ['职业教育本科', '职教本科', '本科']
    }
    
    aliases = level_aliases.get(level, [level])
    
    for item in data:
        edu_level = item.get('education_level', '')
        if any(alias in edu_level for alias in aliases):
            results.append(item)
    
    return results

def search_major(
    query: str,
    level: Optional[str] = None,
    json_path: str = "assets/moe_pdfs_final.json",
    output_path: Optional[str] = None,
    use_fallback: bool = True,
    parse_pdf: bool = True,
    cache_dir: Optional[str] = "temp/pdf_cache"
) -> Dict:
    """
    检索专业教学标准
    
    数据源优先级：
    1. 主数据源：moe_pdfs_final.json
       - 找到专业且有PDF URL → 下载并解析PDF获取完整内容
       - 找到专业但无PDF URL → 返回基本信息
    2. 备用数据源：moe_pdfs_md/{类代码}*.md
       - 仅当在moe_pdfs_final.json中未找到专业时使用
    
    Args:
        query: 专业名称或代码
        level: 教育层次（可选）
        json_path: JSON数据文件路径
        output_path: 输出文件路径（可选）
        use_fallback: 是否使用备用数据源（默认True）
        parse_pdf: 是否解析PDF获取完整内容（默认True）
        cache_dir: PDF缓存目录（默认temp/pdf_cache）
    
    Returns:
        检索结果字典
    """
    # 输入验证
    if not query or not query.strip():
        return {"query": query, "level": level, "total": 0, "results": [], "error": "查询参数不能为空"}
    
    # 确定项目根目录（指向skill/occupation-analysis目录）
    script_dir = Path(__file__).parent
    project_root = script_dir.parent  # scripts/ -> occupation-analysis/
    json_full_path = project_root / json_path
    
    # 如果主数据源不存在，尝试备选数据源
    if not json_full_path.exists():
        json_full_path = project_root / "assets/moe_pdfs_new.json"
    
    # 主数据源检索
    results = []
    data_source = "primary"
    pdf_parsed = False
    
    if json_full_path.exists():
        try:
            data = load_major_data(json_full_path)
            
            # 尝试按代码检索
            if query.isdigit() and len(query) == 6:
                result = search_by_code(data, query)
                if result:
                    results = [result]
            else:
                # 按名称检索
                results = search_by_name(data, query)
            
            # 按层次筛选
            if level and results:
                results = search_by_level(results, level)
            
            # 尝试解析PDF获取完整内容
            if results and parse_pdf:
                for i, result in enumerate(results):
                    pdf_url = result.get('pdf_url', '')
                    major_code = result.get('major_code', '')
                    
                    if pdf_url:
                        print(f"[INFO] 找到PDF URL，尝试解析: {major_code}")
                        pdf_result = parse_pdf_to_result(pdf_url, major_code, project_root, cache_dir)
                        
                        if pdf_result:
                            # 合并PDF解析结果到基本信息
                            result.update(pdf_result)
                            result['pdf_parsed'] = True
                            pdf_parsed = True
                            print(f"[OK] PDF解析成功: {major_code}")
                        else:
                            result['pdf_parsed'] = False
                            result['pdf_parse_error'] = "PDF解析失败"
                            print(f"[WARNING] PDF解析失败，仅返回基本信息: {major_code}")
                    else:
                        result['pdf_parsed'] = False
                        result['pdf_parse_error'] = "无PDF URL"
                
        except (FileNotFoundError, json.JSONDecodeError) as e:
            error_msg = f"主数据源加载失败: {e}"
            print(f"[ERROR] {error_msg}")
            return {
                "query": query,
                "level": level,
                "total": 0,
                "results": [],
                "error": error_msg,
                "data_source": "failed"
            }
    
    # 备用数据源检索（仅当主数据源未找到时）
    if not results and use_fallback:
        print(f"[INFO] 主数据源未找到结果，尝试备用数据源...")
        fallback_results = search_from_fallback(query, level, project_root)
        if fallback_results:
            results = fallback_results
            data_source = "fallback"
    
    # 准备输出
    output = {
        "query": query,
        "level": level,
        "total": len(results),
        "results": results,
        "data_source": data_source
    }
    
    # 保存到文件
    if output_path:
        try:
            output_full_path = project_root / output_path
            output_full_path.parent.mkdir(parents=True, exist_ok=True)
            with open(output_full_path, 'w', encoding='utf-8') as f:
                json.dump(output, f, ensure_ascii=False, indent=2)
            print(f"[OK] 结果已保存到: {output_full_path}")
        except IOError as e:
            print(f"[ERROR] 保存文件失败: {e}")
    
    return output


def parse_pdf_to_result(
    pdf_url: str,
    major_code: str,
    project_root: Path,
    cache_dir: Optional[str] = "temp/pdf_cache"
) -> Optional[Dict]:
    """
    解析PDF获取专业教学标准完整内容
    
    Args:
        pdf_url: PDF文档URL
        major_code: 专业代码
        project_root: 项目根目录
        cache_dir: 缓存目录
    
    Returns:
        解析结果字典，失败返回None
    """
    try:
        # 尝试导入PDF解析模块
        from pdf_parser import fetch_and_parse_pdf, to_dict
        
        # 确定缓存目录
        cache_path = project_root / cache_dir if cache_dir else None
        
        # 调用PDF解析
        standard = fetch_and_parse_pdf(
            url=pdf_url,
            major_code=major_code,
            cache_dir=cache_path,
            use_cache=True
        )
        
        if standard:
            return to_dict(standard)
        
        return None
        
    except ImportError:
        print("[WARNING] PDF解析模块未安装，跳过PDF解析")
        print("[HINT] 请安装依赖: pip install pdfplumber 或 pip install pypdf")
        return None
    except Exception as e:
        print(f"[ERROR] PDF解析异常: {e}")
        return None


def search_from_fallback(
    query: str,
    level: Optional[str],
    project_root: Path
) -> List[Dict]:
    """
    从备用数据源检索专业信息
    
    流程：
    1. 从职业教育专业目录.xlsx获取专业所属类代码
    2. 从moe_pdfs_md目录读取对应的MD文件
    3. 解析MD文件提取专业信息
    
    Args:
        query: 专业名称或代码
        level: 教育层次
        project_root: 项目根目录
    
    Returns:
        检索结果列表
    """
    results = []
    
    try:
        catalog = get_catalog()
        moe_pdfs_md_dir = project_root / "assets/moe_pdfs_md"
        
        # 按专业代码检索
        if query.isdigit() and len(query) == 6:
            category_info = get_category_from_major_code(catalog, query)
            if category_info:
                md_path = get_md_file_path(category_info, moe_pdfs_md_dir)
                if md_path:
                    result = parse_md_file(md_path, query)
                    if result:
                        results.append(result)
        
        # 按专业名称检索
        else:
            from major_catalog_mapper import search_major_in_catalog
            matches = search_major_in_catalog(query, catalog)
            
            for match in matches[:5]:  # 最多返回5个结果
                md_path = get_md_file_path(match, moe_pdfs_md_dir)
                if md_path:
                    result = parse_md_file(md_path, match['专业代码'])
                    if result:
                        results.append(result)
        
        # 按层次筛选（备用数据源中可能没有层次信息，这里简单处理）
        if level and results:
            # 备用数据源中通过专业代码前缀判断层次
            # 中职：7开头，高职：5开头
            level_prefix = {
                '中职': '7',
                '中等职业教育': '7',
                '高职': '5',
                '高等职业教育': '5',
                '职教本科': '2',
                '职业教育本科': '2'
            }
            prefix = level_prefix.get(level)
            if prefix:
                results = [r for r in results if r.get('major_code', '').startswith(prefix)]
    
    except Exception as e:
        print(f"[ERROR] 备用数据源检索失败: {e}")
    
    return results


def parse_md_file(md_path: Path, target_code: str) -> Optional[Dict]:
    """
    解析专业教学标准MD文件
    
    MD文件格式：
    ```
    专业代码 710201
    
    专业名称 计算机应用
    
    基本修业年限 三年
    
    # 职业面向
    ...
    
    # 培养目标定位
    ...
    ```
    
    Args:
        md_path: MD文件路径
        target_code: 目标专业代码
    
    Returns:
        专业信息字典
    """
    if not md_path.exists():
        return None
    
    try:
        with open(md_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        result = {
            'major_code': target_code,
            'major_name': '',
            'education_level': '',
            'career_orientation': '',
            'training_goal': '',
            'raw_text': '',
            'source_file': str(md_path.name),
            'pdf_url': ''
        }
        
        major_section_pattern = rf'专业代码\s+{target_code}\s*\n(.*?)(?=专业代码\s+\d{{6}}|$)'
        major_match = re.search(major_section_pattern, content, re.DOTALL)
        
        if not major_match:
            return result
        
        major_section = major_match.group(1)
        
        result['raw_text'] = re.sub(r'\n{3,}', '\n\n', major_section.strip())
        
        # 提取专业名称
        name_match = re.search(r'专业名称\s+([^\n]+)', major_section)
        if name_match:
            result['major_name'] = name_match.group(1).strip()
        
        # 提取职业面向（# 职业面向 和下一个 # 标题之间的内容）
        orientation_match = re.search(r'# 职业面向\s*\n+(.*?)(?=\n# |\Z)', major_section, re.DOTALL)
        if orientation_match:
            result['career_orientation'] = orientation_match.group(1).strip()
        
        goal_match = re.search(r'培养目标[：:\s]*\n?(.*?)(?=\n\s*\n|\n#|\n二、|\n三、|\n【|$)', major_section, re.DOTALL)
        if goal_match:
            result['training_goal'] = goal_match.group(1).strip()
        
        if target_code.startswith('7'):
            result['education_level'] = '中等职业教育'
        elif target_code.startswith('5'):
            result['education_level'] = '高等职业教育'
        elif target_code.startswith('2'):
            result['education_level'] = '职业教育本科'
        
        return result
        
    except Exception as e:
        print(f"[ERROR] 解析MD文件失败 {md_path}: {e}")
        return None


# ==================== 专业-职业对照表查询 ====================

def get_mapping_sheet_name(level: Optional[str] = None) -> str:
    """
    根据教育层次获取对照表工作表名称
    
    Args:
        level: 教育层次字符串（可选）
        
    Returns:
        工作表名称
    """
    level_lower = (level or '').lower()
    
    if '中职' in level_lower or '中等' in level_lower:
        return '中职-专业职业对照'
    elif '高职' in level_lower or '专科' in level_lower or '高等' in level_lower:
        return '高职专业-专业职业对照'
    elif '本科' in level_lower:
        return '职教本科-专业职业对照'
    else:
        # 默认尝试中职
        return '中职-专业职业对照'


def query_occupation_mapping(
    major_code: str,
    level: Optional[str] = None,
    xlsx_path: str = "assets/专业-职业对照表.xlsx"
) -> List[Dict]:
    """
    从专业-职业对照表查询职业编码
    
    Args:
        major_code: 专业代码（如 740202）
        level: 教育层次（用于确定工作表）
        xlsx_path: 对照表文件路径
        
    Returns:
        职业信息列表 [{"code": "4-03-02-03", "name": "西式烹调师"}, ...]
    """
    if not HAS_OPENPYXL:
        print("[WARNING] openpyxl未安装，无法查询对照表")
        return []
    
    results = []
    
    try:
        from pathlib import Path
        script_dir = Path(__file__).parent
        project_root = script_dir.parent
        full_path = project_root / xlsx_path
        
        if not full_path.exists():
            print(f"[WARNING] 对照表文件不存在: {full_path}")
            return []
        
        wb = openpyxl.load_workbook(full_path)
        
        # 获取工作表名称
        sheet_name = get_mapping_sheet_name(level)
        
        if sheet_name not in wb.sheetnames:
            # 尝试其他工作表
            for name in wb.sheetnames:
                if '中职' in name or '高职' in name or '本科' in name:
                    sheet_name = name
                    break
        
        ws = wb[sheet_name]
        
        # 专业代码可能在第2列（B列），职业编码在第4列（D列），职业名称在第5列（E列）
        # 遍历查找
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 5:
                row_major_code = row[1]  # 专业代码
                
                # 处理专业代码格式（可能是整数或字符串）
                if isinstance(row_major_code, int):
                    row_major_code = str(row_major_code)
                
                # 匹配专业代码
                if row_major_code == str(major_code):
                    occupation_code = row[3]  # 职业编码
                    occupation_name = row[4]  # 职业名称
                    
                    if occupation_code and occupation_name:
                        results.append({
                            "code": str(occupation_code) if isinstance(occupation_code, int) else occupation_code,
                            "name": occupation_name
                        })
        
        if results:
            print(f"[OK] 从对照表查询到 {len(results)} 个职业")
        else:
            print(f"[INFO] 对照表中未找到专业代码 {major_code} 对应的职业")
            
    except Exception as e:
        print(f"[ERROR] 对照表查询失败: {e}")
    
    return results


def main():
    parser = argparse.ArgumentParser(description='检索专业教学标准')
    parser.add_argument('--major', '-m', required=True, help='专业名称或代码')
    parser.add_argument('--level', '-l', help='教育层次（中职/高职/职教本科）')
    parser.add_argument('--json', '-j', default='assets/moe_pdfs_final.json', help='JSON数据文件路径')
    parser.add_argument('--output', '-o', help='输出文件路径')
    parser.add_argument('--no-pdf', action='store_true', help='不解析PDF，仅返回基本信息')
    parser.add_argument('--cache-dir', default='temp/pdf_cache', help='PDF缓存目录')
    parser.add_argument('--include-occupations', action='store_true', 
                        help='同时输出专业-职业对照表中的职业编码')
    
    args = parser.parse_args()
    
    result = search_major(
        query=args.major,
        level=args.level,
        json_path=args.json,
        output_path=args.output,
        parse_pdf=not args.no_pdf,
        cache_dir=args.cache_dir
    )
    
    # 保存完整raw_text到结果中（确保数据完整性）
    if result['results']:
        for item in result['results']:
            # 从PDF缓存读取完整raw_text
            if item.get('major_code') and item.get('pdf_parsed'):
                cache_file = Path(__file__).parent.parent / args.cache_dir / f"{item['major_code']}.json"
                if args.cache_dir and cache_file.exists():
                    try:
                        cache_data = json.loads(cache_file.read_text(encoding='utf-8'))
                        if cache_data.get('raw_text'):
                            item['raw_text'] = cache_data['raw_text']
                    except (json.JSONDecodeError, IOError, OSError):
                        pass
    
    if args.include_occupations and result['results']:
        for item in result['results']:
            major_code = item.get('major_code')
            if major_code:
                occupations = query_occupation_mapping(major_code, args.level)
                
                career_orientation = item.get('career_orientation', '') or item.get('raw_text', '')
                pdf_occ_codes = []
                pdf_occ_names = []
                
                if career_orientation:
                    codes = re.findall(r'[（(]([0-9]+-[0-9]+-[0-9]+-[0-9]+)[）)]', career_orientation)
                    names_match = re.search(r'主要职业类别[（(]代码[）)]\s*(.+?)(?:\n|主要岗位)', career_orientation)
                    
                    if codes:
                        pdf_occ_codes = codes
                    
                    if names_match:
                        names_text = names_match.group(1)
                        name_pattern = r'([^（(]+)[（(]([0-9]+-[0-9]+-[0-9]+-[0-9]+)[）)]'
                        pdf_occ_names = [(re.sub(r'^[、，,\s]+', '', m[0]), m[1]) for m in re.findall(name_pattern, names_text)]
                
                merged_occupations = []
                seen_codes = set()
                
                for occ in occupations:
                    occ_code = occ.get('code', '')
                    occ_name = occ.get('name', '')
                    if occ_code and occ_code not in seen_codes:
                        seen_codes.add(occ_code)
                        merged_occ = {'code': occ_code, 'name': occ_name, 'source': '对照表.xlsx'}
                        
                        pdf_codes_prefix = [c for c in pdf_occ_codes if c[:5] == occ_code[:5]]
                        if pdf_codes_prefix:
                            merged_occ['validation'] = 'prefix_matched'
                        else:
                            merged_occ['validation'] = '对照表仅'
                        
                        merged_occupations.append(merged_occ)
                
                for name, code in pdf_occ_names:
                    if code and code not in seen_codes:
                        seen_codes.add(code)
                        merged_occ = {'code': code, 'name': name.strip(), 'source': 'PDF解析'}
                        
                        mapping_codes = [o.get('code', '') for o in occupations]
                        if any(mc[:5] == code[:5] for mc in mapping_codes):
                            merged_occ['validation'] = 'prefix_matched'
                        else:
                            merged_occ['validation'] = 'PDF新增'
                        
                        merged_occupations.append(merged_occ)
                
                item['occupations_from_mapping'] = merged_occupations
                item['pdf_occupation_codes'] = pdf_occ_codes
                item['occupation_merge_summary'] = {
                    'pdf_count': len(pdf_occ_codes),
                    'mapping_count': len(occupations),
                    'merged_count': len(merged_occupations)
                }
        
        if args.output:
            script_dir = Path(__file__).parent
            project_root = script_dir.parent
            output_full_path = project_root / args.output
            output_full_path.parent.mkdir(parents=True, exist_ok=True)
            with open(output_full_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False, indent=2)
            print(f"[OK] 结果已保存到: {output_full_path}")
    
    print(f"\n检索结果：共找到 {result['total']} 条记录 (数据源: {result['data_source']})")
    for i, item in enumerate(result['results'], 1):
        print(f"\n[{i}] {item['major_name']}")
        print(f"    专业代码: {item['major_code']}")
        print(f"    教育层次: {item['education_level']}")
        print(f"    PDF链接: {item['pdf_url']}")
        if item.get('pdf_parsed'):
            print(f"    PDF解析: ✅ 已解析")
            if item.get('career_orientation'):
                print(f"    职业面向: {item['career_orientation'][:100]}...")
        elif item.get('pdf_parse_error'):
            print(f"    PDF解析: ❌ {item['pdf_parse_error']}")
        
        # 输出职业对照信息
        if item.get('occupations_from_mapping'):
            print(f"    职业对照: {len(item['occupations_from_mapping'])} 个职业")
            for occ in item['occupations_from_mapping']:
                print(f"      - {occ['name']} ({occ['code']})")

if __name__ == '__main__':
    main()
